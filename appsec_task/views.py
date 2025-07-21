from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.shortcuts import redirect
from pentest_task.models import PentestTask, Vulnerability, AffectedURL
from verify_task.models import VerifyTask
from .models import AppSecTask, ShareCostDetails, SecurityException, AppSecTask
from .forms import AppSecTaskForm, ShareCostDetailsForm, SecurityExceptionForm
from django.contrib import messages
from django.db.models import Q, Min, Max, Count
#import, export file
import pandas as pd
import xlsxwriter
from django.http import HttpResponse
from io import BytesIO
import traceback
from datetime import datetime, date
from collections import defaultdict
import openpyxl
# from openpyxl import load_workbook
import calendar
import json
from django.contrib.auth.decorators import login_required
from task_manager.decorators import require_groups
from django.utils.timezone import localtime
from datetime import datetime
import io
import html
from django.contrib.auth.models import User, Group
from urllib.parse import urlparse
from django.core.exceptions import ValidationError
from django.http import HttpResponseBadRequest
import magic  # pip install python-magic
import os
from django.conf import settings
from openpyxl.styles import PatternFill

from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font
from django.core.mail import send_mail
from django.utils import timezone
from django.template.loader import render_to_string
from django.template import Context, Template
from django.utils.html import format_html
from collections import defaultdict
from pentest_task.models import Notification




def safe_str(value):
    return "" if pd.isna(value) or value is None else str(value).strip()

def safe_int(value):
    try:
        return int(value) if not pd.isna(value) and str(value).strip().isdigit() else None
    except ValueError:
        return None


def safe_date(val):
    try:
        if pd.isna(val):
            return None
        if isinstance(val, (datetime, date, pd.Timestamp)):
            return val.date() if isinstance(val, pd.Timestamp) else val
        return pd.to_datetime(str(val)).date()
    except Exception as e:
        # print(f"❌ safe_date error: {e}, val: {val}")
        messages.error(request,f"❌ safe_date error: {e}, val: {val}")
        return None


ALLOWED_EXTENSIONS = ['.docx', '.xlsx', '.xls']
ALLOWED_MIME_TYPES = [
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',  # .docx
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',       # .xlsx
    'application/vnd.ms-excel'                                                 # .xls
]
MAX_FILE_SIZE = 200 * 1024 * 1024  # 100MB

def handle_uploaded_file(file):
    # 1. Kiểm tra dung lượng
    if file.size > MAX_FILE_SIZE:
        raise ValidationError("File size limit 200MB")

    # 2. Kiểm tra đuôi file
    ext = os.path.splitext(file.name)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValidationError("File type is not allow (only docx, xlsx, xls).")

    # 3. Kiểm tra MIME type
    mime = magic.from_buffer(file.read(2048), mime=True)
    file.seek(0)  # reset lại con trỏ để không ảnh hưởng đến xử lý tiếp theo

    if mime not in ALLOWED_MIME_TYPES:
        raise ValidationError(f"MIME type is not allow (only docx, xlsx, xls): {mime}")

    # Nếu hợp lệ: xử lý tiếp...
    return True


@login_required
@require_groups(['Pentester', 'Leader'])
def create_appsec_task(request):
    if request.method == "POST":
        form = AppSecTaskForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('appsec_task:list_appsec_tasks')
    else:
        form = AppSecTaskForm()

    return render(request, 'appsec_task/create_appsec_task.html', {'form': form})


@login_required
@require_groups(['Pentester', 'Leader'])
def edit_appsec_task(request, task_id):
    task = get_object_or_404(AppSecTask, id=task_id)
    if request.method == "POST":
        form = AppSecTaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            return redirect("appsec_task:list_appsec_tasks")
    else:
        form = AppSecTaskForm(instance=task)
    return render(request, "appsec_task/edit_appsec_task.html", {"form": form, "task": task})


@login_required
@require_groups(['Pentester', 'Leader'])
def delete_appsec_task(request, task_id):
    task = get_object_or_404(AppSecTask, id=task_id)
    if request.method == "POST":
        task.delete()
    return redirect('appsec_task:list_appsec_tasks') 


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def view_appsec_task(request, task_id):
    task = get_object_or_404(AppSecTask, id=task_id)
    return render(request, "appsec_task/view_appsec_task.html", {"form": task})


@login_required
@require_groups(['Pentester', 'Leader','Manager'])
def list_appsec_tasks(request):
    # tasks = AppSecTask.objects.all()
    tasks = AppSecTask.objects.prefetch_related("pentest_tasks", "verify_tasks").all()
    share_cost_choices = AppSecTask._meta.get_field('sharecost').choices
    component_choices = AppSecTask._meta.get_field('component').choices
    is_newapp_choices = AppSecTask._meta.get_field('is_newapp').choices 
    is_internet_choices = AppSecTask._meta.get_field('is_internet').choices 
    checklist_type_choices = AppSecTask._meta.get_field('checklist_type').choices 
    status_choices = AppSecTask._meta.get_field('status').choices 
    for task in tasks:
        sync_status(task.id)
    return render(request, 'appsec_task/list_appsec_tasks.html', 
        {'tasks': tasks,
        'share_cost_choices': share_cost_choices,
        'component_choices': component_choices,
        'is_newapp_choices': is_newapp_choices,
        'is_internet_choices': is_internet_choices,
        'checklist_type_choices': checklist_type_choices,
        'status_choices': status_choices,
        })

  
def parse_component_list(raw_value):
    if not raw_value:
        return []
    parts = [p.strip() for p in raw_value.split(",")]
    valid = dict(AppSecTask._meta.get_field('component').choices)
    return [p for p in parts if p in valid]


@login_required
@require_groups(['Pentester', 'Leader'])
def import_appsec_tasks(request):
    if not request.FILES.get("task_file"):
        messages.error(request, "⚠️ Bạn chưa chọn file để upload.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    if request.method == "POST" and request.FILES.get("task_file"):
        file = request.FILES["task_file"]
        try:
            handle_uploaded_file(file)
            # Đọc file từ request.FILES một lần, rồi dùng lại
            uploaded_file = request.FILES["task_file"]
            file_bytes = uploaded_file.read()

            verify_names = set()
            pentest_names = set()
            try:
                xls = pd.ExcelFile(file)

                # Sheet 1: VERIFY TASK
                verify_df = pd.read_excel(xls, sheet_name="Verify Request")
                verify_df.columns = verify_df.columns.str.strip()  # 🔥 Quan trọng

                # Tạo file-like object
                file_buffer = io.BytesIO(file_bytes)

                # Cho pandas đọc từ buffer (nhớ seek về đầu)
                file_buffer.seek(0)
                xls = pd.ExcelFile(file_buffer)
                # Dùng openpyxl để lấy hyperlink
                file_buffer.seek(0)
                wb = openpyxl.load_workbook(file, data_only=True)
                sheet = wb["Verify Request"]
                # Tìm index cột "Sharepoint Link" trong DataFrame
                sharepoint_col_idx = verify_df.columns.get_loc("Sharepoint Link")  # 0-based index

                for idx, row in verify_df.iterrows():
                    try:
                        appsec_name = safe_str(row.get("Task"))
                        if not appsec_name:
                            print("⚠️ Bỏ qua dòng vì không có appsec_name:", row.to_dict())
                            continue
                        # Vì Excel bắt đầu từ hàng 1, còn pandas bỏ qua header (dòng 0), nên cần cộng thêm 2
                        excel_row = idx + 2
                        excel_col = sharepoint_col_idx + 1  # openpyxl dùng 1-based index cho cột

                        cell = sheet.cell(row=excel_row, column=excel_col)
                        link_sharepoint = cell.hyperlink.target if cell.hyperlink else None

                        appsec_task, created = AppSecTask.objects.get_or_create(name=appsec_name, defaults={
                            'description': safe_str(row.get("Description")),
                            'owner': safe_str(row.get("Owner/Requester")),
                            'environment_prod': safe_str(row.get("Domain PROD")),
                            'name_sharepoint': safe_str(row.get("Sharepoint Link")),
                            'link_sharepoint': link_sharepoint,
                            'link_ticket': safe_str(row.get("Ticket")),
                            'mail_loop': safe_str(row.get("Mail loop")),
                            'chat_group': safe_str(row.get("Chat group")),
                            'is_internet': safe_str(row.get("Public Internet/Internal?")),
                            'is_newapp': safe_str(row.get("NewApp/OldApp?")),
                            'checklist_type': safe_str(row.get("Checklist Type")),
                            'sharecost': safe_str(row.get("Share Cost?")),
                            'component': parse_component_list(safe_str(row.get("Component"))),
                            'pentest_vendor': safe_str(row.get("Pentest Vendor")),
                        })

                        if not created:
                            # Cập nhật AppSecTask nếu đã tồn tại
                            appsec_task.description = safe_str(row.get("Description"))
                            appsec_task.owner = safe_str(row.get("Owner/Requester"))
                            appsec_task.environment_prod = safe_str(row.get("Domain PROD"))
                            appsec_task.name_sharepoint = safe_str(row.get("Sharepoint Link"))
                            appsec_task.link_sharepoint = link_sharepoint
                            appsec_task.link_ticket = safe_str(row.get("Ticket"))
                            appsec_task.mail_loop = safe_str(row.get("Mail loop"))
                            appsec_task.chat_group = safe_str(row.get("Chat group"))
                            appsec_task.is_internet = safe_str(row.get("Public Internet/Internal?"))
                            appsec_task.is_newapp = safe_str(row.get("NewApp/OldApp?"))
                            appsec_task.checklist_type = safe_str(row.get("Checklist Type"))
                            appsec_task.sharecost = safe_str(row.get("Share Cost?"))
                            appsec_task.component = parse_component_list(safe_str(row.get("Component")))
                            appsec_task.pentest_vendor = safe_str(row.get("Pentest Vendor"))
                            
                            appsec_task.save()
                            print(f"🔁 Đã cập nhật AppSecTask '{appsec_name}'")
                            messages.warning(request, f"❌ Cập nhật AppsecTask: {appsec_name},link_sharepoint: {link_sharepoint}, row: {row.to_dict()}")
                        verify_names.add(appsec_name)

                        verify_task = VerifyTask.objects.filter(appsec_task=appsec_task).first()
                        if verify_task:
                            verify_task.name = appsec_name
                            verify_task.description = safe_str(row.get("Description"))
                            verify_task.PIC_ISM = safe_str(row.get("PIC ISM"))
                            verify_task.status = safe_str(row.get("Status"))
                            verify_task.start_date = safe_date(row.get("Start date"))
                            verify_task.end_date = safe_date(row.get("Finish date"))
                            verify_task.save()
                            print(f"🔁 Đã cập nhật VerifyTask cho '{appsec_name}'")
                            messages.warning(request, f"❌ Cập nhật VerifyTask: {appsec_name}, row: {row.to_dict()}")
                        else:
                            VerifyTask.objects.create(
                                appsec_task=appsec_task,
                                name=appsec_name,
                                description=safe_str(row.get("Description")),
                                PIC_ISM=safe_str(row.get("PIC ISM")),
                                status=safe_str(row.get("Status")),
                                start_date=safe_date(row.get("Start date")),
                                end_date=safe_date(row.get("Finish date")),
                            )
                            print(f"✅ Tạo VerifyTask cho '{appsec_name}'")

                    except Exception as e:
                        messages.error(request, f"❌ Lỗi tạo/cập nhật VerifyTask: {e}, row: {row.to_dict()}")
                        traceback.print_exc()


                # Sheet 2: PENTEST TASK
                pentest_df = pd.read_excel(xls, sheet_name="Pentest Request")

                # Tạo file-like object
                file_buffer = io.BytesIO(file_bytes)

                # Cho pandas đọc từ buffer (nhớ seek về đầu)
                file_buffer.seek(0)
                xls = pd.ExcelFile(file_buffer)
                # Dùng openpyxl để lấy hyperlink
                file_buffer.seek(0)
                wb = openpyxl.load_workbook(file, data_only=True)
                sheet = wb["Pentest Request"]
                # Tìm index cột "Sharepoint Link" trong DataFrame
                sharepoint_col_idx = pentest_df.columns.get_loc("Sharepoint Link")  # 0-based index
                for idx, row in pentest_df.iterrows():
                    try:
                        appsec_name = safe_str(row.get("Task"))
                        if not appsec_name:
                            print("⚠️ Bỏ qua dòng vì không có appsec_name:", row.to_dict())
                            continue

                        # Vì Excel bắt đầu từ hàng 1, còn pandas bỏ qua header (dòng 0), nên cần cộng thêm 2
                        excel_row = idx + 2
                        excel_col = sharepoint_col_idx + 1  # openpyxl dùng 1-based index cho cột

                        cell = sheet.cell(row=excel_row, column=excel_col)
                        link_sharepoint = cell.hyperlink.target if cell.hyperlink else None
                    
                        appsec_task, created = AppSecTask.objects.get_or_create(name=appsec_name, defaults={
                            'description': safe_str(row.get("Description")),
                            'owner': safe_str(row.get("Owner/Requester")),
                            'environment_prod': safe_str(row.get("Domain PROD")),
                            'name_sharepoint': safe_str(row.get("Sharepoint Link")),
                            'link_sharepoint': link_sharepoint,
                            'link_ticket': safe_str(row.get("Ticket")),
                            'mail_loop': safe_str(row.get("Mail loop")),
                            'chat_group': safe_str(row.get("Chat group")),
                            'is_internet': safe_str(row.get("Public Internet/Internal?")),
                            'is_newapp': safe_str(row.get("NewApp/OldApp?")),
                            'checklist_type': safe_str(row.get("Checklist Type")),
                            'sharecost': safe_str(row.get("Share Cost?")),
                            'component': parse_component_list(safe_str(row.get("Component"))),
                            'pentest_vendor': safe_str(row.get("Pentest Vendor")),
                        })

                        if not created:
                            # Cập nhật AppSecTask nếu đã tồn tại
                            appsec_task.description = safe_str(row.get("Description"))
                            appsec_task.owner = safe_str(row.get("Owner/Requester"))
                            appsec_task.environment_prod = safe_str(row.get("Domain PROD"))
                            appsec_task.name_sharepoint = safe_str(row.get("Sharepoint Link"))
                            # appsec_task.link_sharepoint = link_sharepoint #bỏ ra để chỉ cần update ở verify_task thôi.
                            appsec_task.link_ticket = safe_str(row.get("Ticket"))
                            appsec_task.mail_loop = safe_str(row.get("Mail loop"))
                            appsec_task.chat_group = safe_str(row.get("Chat group"))
                            appsec_task.is_internet = safe_str(row.get("Public Internet/Internal?"))
                            appsec_task.is_newapp = safe_str(row.get("NewApp/OldApp?"))
                            appsec_task.checklist_type = safe_str(row.get("Checklist Type"))
                            appsec_task.sharecost = safe_str(row.get("Share Cost?"))
                            appsec_task.component = parse_component_list(safe_str(row.get("Component")))
                            appsec_task.pentest_vendor = safe_str(row.get("Pentest Vendor"))
                           
                            appsec_task.save()
                            print(f"🔁 Đã cập nhật AppSecTask '{appsec_name}'")

                        pentest_names.add(appsec_name)
                        
                        pentest_task = PentestTask.objects.filter(appsec_task=appsec_task).first()
                        if pentest_task:
                            pentest_task.name = appsec_name
                            pentest_task.description = safe_str(row.get("Description"))
                            pentest_task.environment_test = safe_str(row.get("Domain Test"))
                            pentest_task.status = safe_str(row.get("Status"))
                            pentest_task.ref = safe_str(row.get("REF"))
                            pentest_task.number_of_apis = safe_int(row.get("Number of API/Scope"))
                            pentest_task.effort_working_days = safe_int(row.get("Pentest + Retest Effort (md)/person"))
                            pentest_task.PIC_ISM = safe_str(row.get("PIC ISM"))
                            pentest_task.start_date = safe_date(row.get("Start pentest date"))
                            pentest_task.end_date = safe_date(row.get("Finish pentest date"))
                            pentest_task.start_retest = safe_date(row.get("Start retest date"))
                            pentest_task.end_retest = safe_date(row.get("Finish retest date"))
                            pentest_task.save()
                            print(f"🔁 Đã cập nhật PentestTask cho '{appsec_name}'")
                            messages.warning(request, f"❌ Cập nhật PentestTask: {appsec_name}, row: {row.to_dict()}")
                        else:
                            PentestTask.objects.create(
                                appsec_task=appsec_task,
                                name=appsec_name,
                                description=safe_str(row.get("Description")),
                                environment_test=safe_str(row.get("Domain Test")),
                                status=safe_str(row.get("Status")),
                                ref=safe_str(row.get("REF")),
                                number_of_apis=safe_int(row.get("Number of API/Scope")),
                                effort_working_days=safe_int(row.get("Pentest + Retest Effort (md)/person")),
                                PIC_ISM=safe_str(row.get("PIC ISM")),
                                start_date=safe_date(row.get("Start pentest date")),
                                end_date=safe_date(row.get("Finish pentest date")),
                                start_retest=safe_date(row.get("Start retest date")),
                                end_retest=safe_date(row.get("Finish retest date")),
                                
                            )
                            print(f"✅ Tạo PentestTask cho '{appsec_name}'")
                    except Exception as e:
                        messages.error(request, f"❌ Lỗi tạo/cập nhật PentestTask: {e}, row: {row.to_dict()}")
                        traceback.print_exc()

                
                # Sheet 3: Vulnerability
                
                vulnerabilities = set()
                xls = pd.ExcelFile(file)
                vuln_df = pd.read_excel(xls, sheet_name="Vulnerability")
                for _, row in vuln_df.iterrows():
                    ref=safe_str(row.get("REF")) #giá trị ref của vulnerability
                    ref_prefix = "-".join(ref.split("-")[:-1]) # Lấy phần trước dấu "-" để lấy giá trị ref của pentest_task
                    ref_prefix = safe_str(ref_prefix)  
                    name_vuln_row = safe_str(row.get("Issue Description "))
                    
                    try:
                        pentest_task = PentestTask.objects.get(ref=ref_prefix)
                        if pentest_task:
                            print(f"Vulnerability: Found pentest_task: {pentest_task}")
                            messages.success(request, f"Đã tìm thấy PentestTask với ref: {pentest_task.ref}")
                            # messages.warning(request, f"Đã tìm thấy PentestTask với ref: {ref_prefix}")
                        else:
                            print("Vulnerability: No pentest_task found.")
                            messages.error(request, f"Không tìm thấy PentestTask với ref: {pentest_task.ref}")
                        # ✅ Bỏ qua nếu Vulnerability đã tồn tại với pentest_task và name_vuln
                        if Vulnerability.objects.filter(pentest_task=pentest_task, name_vuln=name_vuln_row).exists():
                            print(f"⚠️ Vulnerability '{name_vuln_row}' đã tồn tại cho task '{ref_prefix}', bỏ qua.")
                            continue

                        Vulnerability.objects.create(
                            pentest_task=pentest_task,
                            ref=ref,
                            name_vuln=safe_str(row.get("Issue Description ")),
                            risk_rating=safe_str(row.get("Risk")),
                            notify_date=safe_date(row.get("Notify")),
                            status=safe_str(row.get("Status")),
                        )
                    except PentestTask.DoesNotExist:
                        messages.warning(request, f"Vulnerability: Không tìm thấy PentestTask với ref: {ref_prefix}")
                        continue
           
                # Sheet 4: Exception
                xls = pd.ExcelFile(file)
                exception_df = pd.read_excel(xls, sheet_name="Exception")
                
                for _, row in exception_df.iterrows():
                    try:
                        task_name = safe_str(row.get("Task"))
                        if not task_name:
                            print("⚠️ Bỏ qua dòng vì không có task_name:", row.to_dict())
                            continue

                        try:
                            appsec_task = AppSecTask.objects.get(name=task_name)
                            messages.success(request, f"✅ Excpetion: Đã tìm thấy AppSecTask: {task_name}")
                        except AppSecTask.DoesNotExist:
                            messages.warning(request, f"❌ Excpetion: Không tìm thấy AppSecTask: {task_name}")
                            continue

                        vulnerability = safe_str(row.get("Vulnerability"))

                        # Kiểm tra nếu đã có exception trùng task + vulnerability
                        exception_obj = SecurityException.objects.filter(
                            appsec_task=appsec_task,
                            vulnerability=vulnerability
                        ).first()
                            # Nếu đã tồn tại → update
                        if exception_obj:
                            exception_obj.risk_level = safe_str(row.get("Risk Level"))
                            exception_obj.status = safe_str(row.get("Status Exception"))
                            exception_obj.exception_date = safe_date(row.get("Exception Expire Date"))
                            exception_obj.exception_create = safe_date(row.get("Exception Create Date"))
                            exception_obj.mail_loop = safe_str(row.get("Mail Loop"))
                            exception_obj.save()
                            print(f"✅ Đã cập nhật Exception: {vulnerability} cho task: {task_name}")
                        else:
                            # Nếu chưa tồn tại → tạo mới
                            SecurityException.objects.create(
                                appsec_task=appsec_task,
                                vulnerability=vulnerability,
                                risk=safe_str(row.get("Risk Level")),
                                status=safe_str(row.get("Status Exception")),
                                exception_date=safe_date(row.get("Exception Expire Date")),
                                exception_create=safe_date(row.get("Exception Create Date")),
                                mail_loop=safe_str(row.get("Mail Loop")),
                            )
                            print(f"✅ Đã tạo mới Exception: {vulnerability} cho task: {task_name}")

                    except Exception as e:
                        messages.error(request, f"❌ Error in Exception import: {e}, row: {row.to_dict()}")
                        continue

   
                messages.success(request, "✅ Tasks, Vulnerability and Exception imported and updated successfully!")

            except Exception as e:
                messages.error(request, f"❌ Lỗi đọc file hoặc xử lý tổng quát: {e}")
                traceback.print_exc()
        except ValidationError as e:
            return HttpResponseBadRequest(str(e))
        return redirect("appsec_task:list_appsec_tasks")


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def export_appsec_tasks(request):
    # sẽ là export everify, all task, sharecost, vuln, exception
    # Lấy tất cả verify_task + pentest_task
    verify_tasks = VerifyTask.objects.select_related("appsec_task").all()
    pentest_tasks = PentestTask.objects.select_related("appsec_task").all()
    vulnerabilities = Vulnerability.objects.select_related("pentest_task").all()  
    security_exceptions = SecurityException.objects.select_related("appsec_task").all()

    # Sheet 1: Verify Request 2025 (All)
    verify_data = []
    for task in verify_tasks:
        verify_data.append({
            "Task": task.appsec_task.name if task.appsec_task else "",
            "Description": task.description,
            "Owner/Requester": task.appsec_task.owner if task.appsec_task else "",
            "Domain PROD": task.appsec_task.environment_prod if task.appsec_task else "",
            "PIC ISM": task.PIC_ISM,
            "Status": task.status,
            "Start date": task.start_date,
            "Finish date": task.end_date,

            "Sharepoint Link": task.appsec_task.link_sharepoint if task.appsec_task else "",
            "Sharepoint Name": task.appsec_task.name_sharepoint if task.appsec_task else "",
            "Ticket": task.appsec_task.link_ticket if task.appsec_task else "",
            "Mail loop": task.appsec_task.mail_loop if task.appsec_task else "",
            "Chat group": task.appsec_task.chat_group if task.appsec_task else "",
            "Public Internet/Internal?": task.appsec_task.is_internet if task.appsec_task else "",
            "NewApp/OldApp?": task.appsec_task.is_newapp if task.appsec_task else "",
            "Checklist Type": task.appsec_task.checklist_type if task.appsec_task else "",
            "Share Cost?": task.appsec_task.sharecost,
            "Component": ", ".join(task.appsec_task.component) if task.appsec_task and task.appsec_task.component else "",
            "Pentest Vendor": task.appsec_task.pentest_vendor,
        })
    verify_df = pd.DataFrame(verify_data)
    # Chuyển sang DataFrame nhưng không export cột 'link_raw'
    verify_df_display = verify_df.drop(columns=["Sharepoint Name"])


    # Sheet 2: Pentest Request 2025
    pentest_data = []
    for task in pentest_tasks:
        pentest_data.append({
            "Task": task.appsec_task.name if task.appsec_task else "",
            "Description": task.description,
            "Owner/Requester": task.appsec_task.owner if task.appsec_task else "",
            "Domain PROD": task.appsec_task.environment_prod if task.appsec_task else "",
            "Domain Test": task.environment_test,
            
            "PIC ISM": task.PIC_ISM,
            "Status": task.status,
            "Start pentest date": task.start_date,
            "Finish pentest date": task.end_date,
            "Start retest date": task.start_retest,
            "Finish retest date": task.end_retest,
            "REF": task.ref,
            "Number of API/Scope": task.number_of_apis,
            "Pentest + Retest Effort (md)/person": task.effort_working_days,
            
            "Sharepoint Link": task.appsec_task.link_sharepoint if task.appsec_task else "",
            "Sharepoint Name": task.appsec_task.name_sharepoint if task.appsec_task else "",
            "Ticket": task.appsec_task.link_ticket if task.appsec_task else "",
            "Mail loop": task.appsec_task.mail_loop if task.appsec_task else "",
            "Chat group": task.appsec_task.chat_group if task.appsec_task else "",
            "Public Internet/Internal?": task.appsec_task.is_internet if task.appsec_task else "",
            "NewApp/OldApp?": task.appsec_task.is_newapp if task.appsec_task else "",
            "Checklist Type": task.appsec_task.checklist_type if task.appsec_task else "",
            "Component": ", ".join(task.appsec_task.component) if task.appsec_task and task.appsec_task.component else "",
            "Pentest Vendor": task.appsec_task.pentest_vendor,
            "Share Cost?": task.appsec_task.sharecost,
            
        })
    pentest_df = pd.DataFrame(pentest_data)
    pentest_df_display = pentest_df.drop(columns=["Sharepoint Name"])
    # Sheet 3: Vulnerability
    vuln_data = []
    for vuln in vulnerabilities:
        vuln_data.append({
            "Task": vuln.pentest_task.name if vuln.pentest_task else "",
            "Domain Test": vuln.pentest_task.environment_test if vuln.pentest_task else "",
            "Issue Description": vuln.name_vuln,
            "REF": vuln.ref,
            "Risk": vuln.risk_rating,
            "Notify":vuln.notify_date,
            "Status": vuln.status,
            "PIC": vuln.pentest_task.PIC_ISM,
            "Component": vuln.pentest_task.appsec_task.component,
            "Public Internet/Internal?": vuln.pentest_task.appsec_task.is_internet if vuln.pentest_task.appsec_task else "",
            "NewApp/OldApp?": vuln.pentest_task.appsec_task.is_newapp if vuln.pentest_task.appsec_task else "",
            "Checklist Type": vuln.pentest_task.appsec_task.checklist_type if vuln.pentest_task.appsec_task else "",
            
        })
    vuln_df = pd.DataFrame(vuln_data)

    # Sheet 4: Exception
    exception_data = []
    for exception in security_exceptions:
        exception_data.append({
            "Task": exception.appsec_task.name if exception.appsec_task else "",
            "Application/Domain name": exception.appsec_task.environment_prod if exception.appsec_task else "",
            "Vulnerability": exception.vulnerability,
            "Risk Level": exception.risk_level,
            "Status Exception": exception.status,
            "Exception Create Date":exception.exception_create,
            "Exception Expire Date":exception.exception_date,
            "Owner/PIC":exception.appsec_task.owner,

            "PIC ISM": exception.appsec_task.PIC_ISM,
            "Mail Loop": exception.mail_loop,
            "Sharepoint link": exception.appsec_task.link_sharepoint if exception.appsec_task else "",
        })
    exception_df = pd.DataFrame(exception_data)



    def escape_excel_formula(s):
        if not isinstance(s, str):
            return ""
        return s.replace('"', '""')  # Excel escape dấu "

    
    def is_valid_url(url):
        try:
            result = urlparse(url)
            return result.scheme in ["http", "https", "ftp", "mailto", "file"]
        except:
            return False

    def input_link(name_df, writer, sheet="Verify Request" ):

        # Chỉ ghi các cột bạn muốn (loại bỏ 'Sharepoint Name')
        export_df = name_df.drop(columns=["Sharepoint Name"])
        export_df.to_excel(writer, sheet_name=sheet, index=False)

        worksheet = writer.sheets[sheet]
        col_link = export_df.columns.get_loc("Sharepoint Link")

        for row in range(len(export_df)):
            link = name_df.iloc[row]["Sharepoint Link"]      # vẫn lấy từ verify_df gốc
            name = name_df.iloc[row]["Sharepoint Name"]      # dùng làm text hiển thị

            if pd.notna(link) and str(link).strip() != "":
                worksheet.write_url(
                    row + 1,
                    col_link,
                    str(link),
                    string=str(name)
                )

    output = BytesIO()

    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        verify_df.to_excel(writer, sheet_name="Verify Request", index=False)
        input_link(verify_df, writer, "Verify Request")
        
        pentest_df.to_excel(writer, sheet_name="Pentest Request", index=False)
        input_link(pentest_df, writer, "Pentest Request")
        
        vuln_df.to_excel(writer, sheet_name="Vulnerability", index=False) 
        exception_df.to_excel(writer, sheet_name="Exception", index=False) 
        

    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    today_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"ISM AppSec FollowUp_{today_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



def sync_status(appsec_task_id):
    appsec_task = get_object_or_404(AppSecTask, id=appsec_task_id)
    pentest_task = appsec_task.pentest_tasks.first()
    verify_task = appsec_task.verify_tasks.first()

    def is_active(status):
        return status not in ["Not Start", "Done", "Cancel", "Interrupt"]

    def is_not_active(status):
        return status in ["In Progress", "Reported", "Retest"]

    # ==== Đồng bộ status ====
    if pentest_task and verify_task:
        pt_status = pentest_task.status
        vt_status = verify_task.status

        if pt_status == "Cancel" or vt_status == "Cancel":
            appsec_task.status = "Cancel"
        elif pt_status == "Interrupt" or vt_status == "Interrupt":
            appsec_task.status = "Interrupt"
        elif pt_status == "Done" and vt_status == "Done":
            appsec_task.status = "Done"
        elif pt_status == "Not Start" and vt_status == "Not Start":
            appsec_task.status = "Not Start"
        elif is_not_active(pt_status) or is_not_active(vt_status):
            appsec_task.status = "In Progress"
            
        else: 
            appsec_task.status = "In Progress"
        # Đồng bộ ngày
        start_dates = [d for d in [verify_task.start_date, pentest_task.start_date, pentest_task.start_retest] if d]
        end_dates = [d for d in [verify_task.end_date, pentest_task.end_date, pentest_task.end_retest] if d]

        if start_dates:
            appsec_task.start_date = min(start_dates)
        if end_dates:
            appsec_task.end_date = max(end_dates)

       

    elif pentest_task and not verify_task:
        appsec_task.status = pentest_task.status
        appsec_task.start_date = pentest_task.start_date
        appsec_task.end_date = pentest_task.end_date
        

    elif verify_task and not pentest_task:
        appsec_task.status = verify_task.status
        appsec_task.start_date = verify_task.start_date
        appsec_task.end_date = verify_task.end_date
        
    
    # Đồng bộ PIC_ISM, loại bỏ trùng lặp (không phân biệt hoa thường) và giữ nguyên định dạng ban đầu
    def normalize_pic(pic_string):
        if not pic_string:
            return []
        return [p.strip() for p in pic_string.split(",") if p.strip()]

    pic_dict = {}
    pic_set = set()

    # Ưu tiên định dạng đầu tiên nếu trùng tên (case-insensitive)
    def add_pics(pic_string):
        for p in normalize_pic(pic_string):
            key = p.lower()
            if key not in pic_dict:
                pic_dict[key] = p  # giữ nguyên định dạng gốc lần đầu tiên xuất hiện

    if pentest_task and pentest_task.PIC_ISM:
        add_pics(pentest_task.PIC_ISM)

    if verify_task and verify_task.PIC_ISM:
        add_pics(verify_task.PIC_ISM)

    appsec_task.PIC_ISM = ", ".join(sorted(pic_dict.values())) if pic_dict else None

    appsec_task.pentest_task = pentest_task
    appsec_task.verify_task = verify_task
    appsec_task.save()


@login_required
@require_groups(['Leader', 'Manager'])
def add_sharecost(request, appsec_task_id):
    appsec_task = get_object_or_404(AppSecTask, id=appsec_task_id)
    if request.method == 'POST':
        form = ShareCostDetailsForm(request.POST)
        if form.is_valid():
            share_cost = form.save(commit=False)
            share_cost.appsec_task = appsec_task
            share_cost.pentest_vendor = appsec_task.pentest_vendor

            share_cost.save()
            messages.success(request, "Sharecost created successfully.")
            return redirect('appsec_task:list_sharecost')
    else:
        form = ShareCostDetailsForm()
    return render(request, 'appsec_task/add_share_cost.html', {'form': form, 'appsec_task': appsec_task})


@login_required
@require_groups(['Leader', 'Manager'])
def edit_sharecost(request, appsec_task_id, sharecost_id):
    appsec_task = get_object_or_404(AppSecTask, id=appsec_task_id)
    sharecost = get_object_or_404(ShareCostDetails, id=sharecost_id, appsec_task=appsec_task)

    if request.method == "POST":
        form = ShareCostDetailsForm(request.POST, instance=sharecost)
        if form.is_valid():
            share_cost = form.save(commit=False)
            share_cost.appsec_task = appsec_task
            share_cost.pentest_vendor = appsec_task.pentest_vendor

            share_cost.save()
            
            return redirect("appsec_task:list_sharecost")
    else:
        form = ShareCostDetailsForm(instance=sharecost)
    return render(request, "appsec_task/add_share_cost.html", {"form": form, "task": sharecost, "appsec_task":appsec_task})


@login_required
@require_groups(['Leader', 'Manager'])
def view_sharecost(request, appsec_task_id, sharecost_id):
    sharecost = get_object_or_404(ShareCostDetails, id=sharecost_id)
    if sharecost.appsec_task.id != appsec_task_id:
        raise PermissionDenied("Sharecost không thuộc về Appsec Task này.")

    appsec_task = sharecost.appsec_task
    return render(request, 'appsec_task/view_share_cost.html', {'form': sharecost, "appsec_task":appsec_task})


@login_required
@require_groups(['Leader', 'Manager'])
def delete_sharecost(request, appsec_task_id, sharecost_id):
    task = get_object_or_404(ShareCostDetails, id=sharecost_id)
    if task.appsec_task.id != appsec_task_id:
        raise PermissionDenied("Sharecost không thuộc về Appsec Task này.")

    if request.method == "POST":
        task.delete()
    return redirect('appsec_task:list_appsec_tasks') 


@login_required
@require_groups(['Leader', 'Manager'])
def list_sharecost(request):
    tasks = ShareCostDetails.objects.all()
    share_cost_choices = AppSecTask._meta.get_field('sharecost').choices
    pay_status_choices = ShareCostDetails._meta.get_field('pay_status').choices
    
    monthly_totals = defaultdict(lambda: {'cost_mm': 0, 'cost_dolla': 0})
    quarterly_totals = defaultdict(lambda: {'cost_mm': 0, 'cost_dolla': 0})

    for task in tasks:
        if task.month_pay:
            try:
                # Parse chuỗi "1/2025"
                month_str, year_str = task.month_pay.strip().split("/")
                month = int(month_str)
                year = int(year_str)
                month_key = f"{year}-{month:02d}"
                quarter = (month - 1) // 3 + 1
                quarter_key = f"Q{quarter}-{year}"
            except ValueError:
                month_key = "Invalid"
                quarter_key = "Invalid"
        else:
            month_key = "N/A"
            quarter_key = "N/A"

        # Gộp tổng
        monthly_totals[month_key]['cost_mm'] += task.cost_mm or 0
        monthly_totals[month_key]['cost_dolla'] += task.cost_dolla or 0
        quarterly_totals[quarter_key]['cost_mm'] += task.cost_mm or 0
        quarterly_totals[quarter_key]['cost_dolla'] += task.cost_dolla or 0


    return render(request, 'appsec_task/list_sharecost.html', 
        {'tasks': tasks,
        'share_cost_choices': share_cost_choices,
        'pay_status_choices': pay_status_choices,
        'monthly_totals': dict(monthly_totals),
        'quarterly_totals': dict(quarterly_totals),
        
        })


@login_required
@require_groups(['Leader', 'Manager'])
def export_sharecost_excel(request):
    # Lấy các filter hiện tại từ query params
    queryset = ShareCostDetails.objects.all()

    # Tạo workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ShareCost"

    # Header
    headers = ["Task", "Sharecost?", "Pentest Vendor", "Project Code", "Owner", "Cost (MM)",  "Cost Dolla", "Month Pay", "Pay Status", "Note",]
    ws.append(headers)

    # Data rows
    for obj in queryset:
        ws.append([
            obj.appsec_task.name if obj.appsec_task else "",  # lấy tên task
            obj.appsec_task.sharecost,
            obj.appsec_task.pentest_vendor,
            obj.project_code,
            obj.owner,
            float(obj.cost_mm) if obj.cost_mm else 0,
            float(obj.cost_dolla) if obj.cost_dolla else 0,
            obj.month_pay,
            obj.pay_status,
            obj.note,
            
        ])

    # Trả file về client
    today_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"sharecost_report_{today_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@login_required
@require_groups(['Leader', 'Manager'])
def import_sharecost(request):
    if not request.FILES.get("task_file"):
        messages.error(request, "⚠️ Bạn chưa chọn file để upload.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    if request.method == "POST" and request.FILES.get("task_file"):
        file = request.FILES["task_file"]
        try:
            handle_uploaded_file(file)
            xls = pd.ExcelFile(file)
            sharecost_df = pd.read_excel(xls, sheet_name="ShareCost")

            for _, row in sharecost_df.iterrows():
                try:
                    task_name = safe_str(row.get("Task"))
                    if not task_name:
                        print("⚠️ Bỏ qua dòng vì không có task_name:", row.to_dict())
                        continue

                    try:
                        appsec_task = AppSecTask.objects.get(name=task_name)
                        
                    except AppSecTask.DoesNotExist:
                        messages.warning(request, f"❌ Sharecost: Không tìm thấy AppSecTask: {task_name}")
                        continue

                    sharecost_obj = ShareCostDetails.objects.filter(appsec_task=appsec_task).first()
                    
                    if sharecost_obj:
                        sharecost_obj.project_code = safe_str(row.get("Project Code"))
                        sharecost_obj.owner = safe_str(row.get("Owner"))
                        sharecost_obj.cost_mm = row.get("Cost (MM)")
                        sharecost_obj.cost_dolla = row.get("Cost Dolla")
                        sharecost_obj.month_pay = safe_str(row.get("Month Pay"))
                        sharecost_obj.pay_status = safe_str(row.get("Pay Status"))
                        sharecost_obj.note = safe_str(row.get("Note"))
                        sharecost_obj.save()
                        messages.warning(request, f"✅ Sharecost: updated Sharecost for AppSecTask: {task_name}")
                        print(f"✅ Sharecost: updated Sharecost for AppSecTask: {task_name}")
                    else:
                        ShareCostDetails.objects.create(
                            appsec_task=appsec_task,
                            project_code=safe_str(row.get("Project Code")),
                            owner=safe_str(row.get("Owner")),
                            cost_mm=row.get("Cost (MM)"),
                            cost_dolla=row.get("Cost Dolla"),
                            month_pay=safe_str(row.get("Month Pay")),
                            pay_status=safe_str(row.get("Pay Status")),
                            note=safe_str(row.get("Note")),
                        )
                        messages.success(request, f"✅ Sharecost: created Sharecost for AppSecTask: {task_name}")
                        print(f"✅ Sharecost: created Sharecost for AppSecTask: {task_name}")

                except Exception as e:
                    print(f"❌ Error in Sharecost import: {e}, row: {row.to_dict()}")
                    messages.error(request, f"❌ Error in Sharecost import: {e}, row: {row.to_dict()}")
                    continue

        except Exception as e:
            traceback.print_exc()
            messages.error(request, f"❌ Error while processing import sharecost data: {e}")
            return HttpResponse(f"Lỗi xử lý file: {e}", status=500)

        return redirect("appsec_task:list_sharecost")

    return HttpResponseBadRequest("❌ No file uploaded.")


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def all_exceptions(request):
    exceptions = SecurityException.objects.all()
    status_choices = SecurityException._meta.get_field('status').choices 
    exploitability_level_choices = SecurityException._meta.get_field('exploitability_level').choices 
    impact_level_choices = SecurityException._meta.get_field('impact_level').choices 
    risk_level_choices = SecurityException._meta.get_field('risk_level').choices 
    return render(request, 'appsec_task/all_exceptions.html', {
        'exceptions': exceptions,
        'status_choices': status_choices,
        'exploitability_level_choices': exploitability_level_choices,
        'impact_level_choices': impact_level_choices,
        'risk_level_choices': risk_level_choices,
        })


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def exception_list(request, appsec_task_id):
    task = get_object_or_404(AppSecTask, id=appsec_task_id)
    exceptions = SecurityException.objects.filter(appsec_task=task)
    status_choices = SecurityException._meta.get_field('status').choices 
    exploitability_level_choices = SecurityException._meta.get_field('exploitability_level').choices 
    impact_level_choices = SecurityException._meta.get_field('impact_level').choices 
    risk_level_choices = SecurityException._meta.get_field('risk_level').choices 
    return render(request, 'appsec_task/list_exception.html', {
        'exceptions': exceptions,
        'appsec_task': task,
        'status_choices': status_choices,
        'exploitability_level_choices': exploitability_level_choices,
        'impact_level_choices': impact_level_choices,
        'risk_level_choices': risk_level_choices,
        })


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def exception_create(request, appsec_task_id):
    task = get_object_or_404(AppSecTask, id=appsec_task_id)
    if request.method == 'POST':
        form = SecurityExceptionForm(request.POST)
        if form.is_valid():
            exception = form.save(commit=False)
            exception.appsec_task = task
            exception.save()
            messages.success(request, "Exception created successfully.")
            return redirect('appsec_task:exception_list', appsec_task_id=task.id)
    else:
        form = SecurityExceptionForm()
    return render(request, 'appsec_task/create_exception.html', {'form': form, 'appsec_task': task})


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def exception_edit(request, appsec_task_id, pk):
    task = get_object_or_404(AppSecTask, id=appsec_task_id)
    exception = get_object_or_404(SecurityException, pk=pk, appsec_task=task)
    if request.method == 'POST':
        form = SecurityExceptionForm(request.POST, instance=exception)
        if form.is_valid():
            exception = form.save(commit=False)
            exception.appsec_task = task
            exception.save()
            messages.success(request, "Exception created successfully.")
            return redirect('appsec_task:exception_list', appsec_task_id=task.id)

    else:
        form = SecurityExceptionForm(instance=exception)
    return render(request, 'appsec_task/create_exception.html', {'form': form, 'appsec_task': task})


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def exception_delete(request, appsec_task_id, pk):
    task = get_object_or_404(AppSecTask, id=appsec_task_id)
    exception = get_object_or_404(SecurityException, pk=pk, appsec_task=task)
    if request.method == 'POST':
        exception.delete()
        messages.success(request, "Exception deleted successfully.")
        return redirect('appsec_task:exception_list', appsec_task_id=task.id)
    return render(request, 'appsec_task/exception_confirm_delete.html', {'exception': exception, 'appsec_task': task})


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def exception_detail(request, appsec_task_id, pk):
    task = get_object_or_404(AppSecTask, id=appsec_task_id)
    exception = get_object_or_404(SecurityException, pk=pk, appsec_task=task)
    return render(request, 'appsec_task/view_exception.html', {'form': exception, 'appsec_task': task})


#dashboard hiển thị:
def get_vuln_stats(selected_year):
    current_year = selected_year
    vuln_stats = {
        'labels': [],
        'affected_apps': [],
        'critical_open': [],
        'critical_close': [],
        'high_open': [],
        'high_close': [],
        'medium_open': [],
        'medium_close': [],
        'total_all': [],
        'total_closed': [],
    }

    for month in range(1, 13):
        vuln_stats['labels'].append(f'{month}')
        start_date = date(current_year, month, 1)
        last_day = calendar.monthrange(current_year, month)[1]
        end_date = date(current_year, month, last_day)

        base_filter = {
            'notify_date__gte': start_date,
            'notify_date__lte': end_date,
            'notify_date__isnull': False
        }

        tasks = Vulnerability.objects.filter(**base_filter).only('pentest_task_id')
        unique_task_ids = set(str(v.pentest_task_id) for v in tasks if v.pentest_task_id)
        vuln_stats['affected_apps'].append(len(unique_task_ids))

        def count_vulns(risk, status):
            return Vulnerability.objects.filter(
                **base_filter,
                risk_rating__iexact=risk,
                status__iexact=status
            ).count()

        total_all = (
            count_vulns('Critical', 'Closed') +
            count_vulns('High', 'Closed') +
            count_vulns('Medium', 'Closed') +
            count_vulns('Critical', 'Open') +
            count_vulns('High', 'Open') +
            count_vulns('Medium', 'Open')
        )
        total_closed = (
            count_vulns('Critical', 'Closed') +
            count_vulns('High', 'Closed') +
            count_vulns('Medium', 'Closed')
        )

        vuln_stats['critical_open'].append(count_vulns('Critical', 'Open'))
        vuln_stats['critical_close'].append(count_vulns('Critical', 'Closed'))
        vuln_stats['high_open'].append(count_vulns('High', 'Open'))
        vuln_stats['high_close'].append(count_vulns('High', 'Closed'))
        vuln_stats['medium_open'].append(count_vulns('Medium', 'Open'))
        vuln_stats['medium_close'].append(count_vulns('Medium', 'Closed'))
        vuln_stats['total_all'].append(total_all)
        vuln_stats['total_closed'].append(total_closed)

    return vuln_stats


def get_affected_url_stats_by_month(selected_year):
    current_year = selected_year
    stats = {
        'labels': [],
        'affected_apps': [],
        'critical_open': [],
        'critical_close': [],
        'high_open': [],
        'high_close': [],
        'medium_open': [],
        'medium_close': [],
        'total_all': [],
        'total_closed': [],
    }

    for month in range(1, 13):
        stats['labels'].append(f'{month}')
        start_date = date(current_year, month, 1)
        last_day = calendar.monthrange(current_year, month)[1]
        end_date = date(current_year, month, last_day)

        base_filter = Q(notify_date__gte=start_date, notify_date__lte=end_date, notify_date__isnull=False)

        
        # Tính affected apps theo pentest_task_id trong AffectedURL
        task_ids = AffectedURL.objects.filter(
            vulnerability__notify_date__gte=start_date,
            vulnerability__notify_date__lte=end_date,
            vulnerability__notify_date__isnull=False,
        ).values_list('vulnerability__pentest_task_id', flat=True).distinct()

        stats['affected_apps'].append(len(set(task_ids)))


        def count_vulns(risk, status):
            return AffectedURL.objects.filter(
                vulnerability__notify_date__gte=start_date,
                vulnerability__notify_date__lte=end_date,
                vulnerability__notify_date__isnull=False,
                vulnerability__risk_rating__iexact=risk,
                status__iexact=status
            ).count()

        critical_open = count_vulns('Critical', 'Open')
        critical_close = count_vulns('Critical', 'Closed')
        high_open = count_vulns('High', 'Open')
        high_close = count_vulns('High', 'Closed')
        medium_open = count_vulns('Medium', 'Open')
        medium_close = count_vulns('Medium', 'Closed')

        total_all = critical_open + critical_close + high_open + high_close + medium_open + medium_close
        total_closed = critical_close + high_close + medium_close

        stats['critical_open'].append(critical_open)
        stats['critical_close'].append(critical_close)
        stats['high_open'].append(high_open)
        stats['high_close'].append(high_close)
        stats['medium_open'].append(medium_open)
        stats['medium_close'].append(medium_close)
        stats['total_all'].append(total_all)
        stats['total_closed'].append(total_closed)

    return stats



def get_affected_url_stats(selected_year):
    affected_urls = AffectedURL.objects.select_related('vulnerability').all()

    monthly_totals = defaultdict(int)
    monthly_closed = defaultdict(int)

    for au in affected_urls:
        vuln = au.vulnerability
        if not vuln or not vuln.notify_date:
            continue
        if vuln.notify_date.year != selected_year:
            continue

        month = vuln.notify_date.month
        monthly_totals[month] += 1

        if vuln.status and vuln.status.lower() == 'closed':
            monthly_closed[month] += 1

    affected_labels = list(range(1, 13))
    affected_total = [monthly_totals.get(month, 0) for month in affected_labels]
    affected_closed = [monthly_closed.get(month, 0) for month in affected_labels]

    return {
        "affected_labels_json": json.dumps(affected_labels),
        "affected_total_json": json.dumps(affected_total),
        "affected_closed_json": json.dumps(affected_closed),

        "affected_labels": affected_labels,
        "affected_total": affected_total,
        "affected_closed": affected_closed,
    }


def get_exception_stats(selected_year):
    exceptions = SecurityException.objects.filter(
        exception_create__isnull=False,
        exception_create__year=selected_year
    )

    monthly_total = defaultdict(int)
    monthly_closed = defaultdict(int)

    for exception in exceptions:
        month = exception.exception_create.month
        monthly_total[month] += 1
        if exception.status.lower() == 'closed':
            monthly_closed[month] += 1

    labels = list(range(1, 13))  # 1 đến 12 tháng
    total_data = [monthly_total.get(month, 0) for month in labels]
    closed_data = [monthly_closed.get(month, 0) for month in labels]

    return {
        "exception_labels_json": json.dumps(labels),
        "exception_total_json": json.dumps(total_data),
        "exception_closed_json": json.dumps(closed_data),

        "exception_labels": labels,
        "exception_total": total_data,
        "exception_closed": closed_data,
    }

    
def task_timeline(current_year):
    pentest_all = PentestTask.objects.filter(start_date__year=current_year)
    retest_all = PentestTask.objects.filter(start_retest__year=current_year)
    verify_all = VerifyTask.objects.filter(start_date__year=current_year)

    def serialize_task(task, task_type):
        data = {
            "name": task.name,
            "pic": task.PIC_ISM or "",
            "start": task.start_date.strftime("%Y-%m-%d") if task.start_date else "",
            "end": task.end_date.strftime("%Y-%m-%d") if task.end_date else "",
            "type": task_type,
            "status": task.status if hasattr(task, "status") else "", 
    
        }

        if task_type == "retest":
            data["start"] = task.start_retest.strftime("%Y-%m-%d") if task.start_retest else ""
            data["end"] = task.end_retest.strftime("%Y-%m-%d") if task.end_retest else ""

        return data

    pentest_retest_tasks = [
        *[serialize_task(task, "pentest") for task in pentest_all],
        *[serialize_task(task, "retest") for task in retest_all],
    ]
    verify_tasks = [serialize_task(task, "verify") for task in verify_all]

     # Lấy tất cả PIC từ User
    # user_choices = [(user.username, user.username) for user in User.objects.all()]
    pentester_group = Group.objects.get(name="Pentester")
    pentesters = User.objects.filter(groups=pentester_group)

    return {
        "pentest_tasks_json": json.dumps(pentest_retest_tasks),
        "verify_tasks_json": json.dumps(verify_tasks),
        "user_choices": pentesters, 
    }

def top_10_vuln_apis(year):
    top_10_common_vulns = (
        Vulnerability.objects
        .filter(pentest_task__start_date__year=year)
        .values('name_vuln')
        .annotate(vuln_count=Count('id'))
        .order_by('-vuln_count')[:10]
    )

    top_10_by_api_count = (
        Vulnerability.objects
        .filter(pentest_task__start_date__year=year)
        .values('name_vuln')
        .annotate(api_count=Count('affected_urls'))
        .order_by('-api_count')[:10]
    )

    return {
        'top_common': list(top_10_common_vulns),
        'top_api': list(top_10_by_api_count),
    }


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def dashboard(request):
    # current_year = datetime.now().year
    year_param = request.GET.get("year")  # Lấy giá trị từ URL: ?year=2024
    try:
        current_year = int(year_param)
    except (TypeError, ValueError):
        current_year = datetime.now().year

    # Tạo danh sách năm từ 2020 đến hiện tại (có thể tùy chỉnh theo dữ liệu bạn có)
    years = list(range(2023, datetime.now().year + 1))


    pentest_counts = defaultdict(int)
    verify_counts = defaultdict(int)

    pentest_tasks = PentestTask.objects.filter(start_date__year=current_year)
    verify_tasks = VerifyTask.objects.filter(start_date__year=current_year)

    for task in pentest_tasks:
        if task.start_date and hasattr(task.start_date, 'month'):
            month = f"{task.start_date.month:02d}"
            pentest_counts[month] += 1

    for task in verify_tasks:
        if task.start_date and hasattr(task.start_date, 'month'):
            month = f"{task.start_date.month:02d}"
            verify_counts[month] += 1

    # List tháng cố định
    months = [f"{i:02d}" for i in range(1, 13)]
    task_stats = [
        {
            "month": month,
            "pentest": pentest_counts.get(month, 0),
            "verify": verify_counts.get(month, 0),
        }
        for month in months
    ]
    # lấy các info về Critical, High, Medium với status là Open và Closed
    vuln_stats = get_vuln_stats(current_year)
    
    # lấy info tổng lỗi Critical+High+Meidum với status là Open và Closed
    # combined_vuln_stats = list(zip(vuln_stats['labels'], vuln_stats['total_all'], vuln_stats['total_closed']))
    combined_vuln_stats = list(zip(
        vuln_stats['labels'],
        vuln_stats['total_all'],
        vuln_stats['total_closed']
    ))

    affected_url_stats = get_affected_url_stats(current_year)
    affected_stats_combined = list(zip(
        affected_url_stats["affected_labels"],
        affected_url_stats["affected_total"],
        affected_url_stats["affected_closed"]
    ))
    
    exception_stats = get_exception_stats(current_year)
    exception_stats_combined = list(zip(
        exception_stats["exception_labels"],
        exception_stats["exception_total"],
        exception_stats["exception_closed"]
    ))
    timeline_stats = task_timeline(current_year)
        
    top_vulns = top_10_vuln_apis(current_year)

    apis_starts = get_affected_url_stats_by_month(current_year)

    context = {
        "years": years,
        "selected_year": current_year,
        "task_stats": task_stats,
        "pentest_labels": months,
        "pentest_data": [item["pentest"] for item in task_stats],
        "verify_data": [item["verify"] for item in task_stats],
        # hiẻn thị trong table
        "vuln_stats_combined": combined_vuln_stats,

        #hiển thị trong vulnChart
        "vuln_stats": vuln_stats,

        #hiển thị trong vulnChart2, sẽ hiện info sát với table hơn
        "vuln_labels_json": json.dumps(vuln_stats['labels']),
        "vuln_total_json": json.dumps(vuln_stats['total_all']),
        "vuln_closed_json": json.dumps(vuln_stats['total_closed']),

        "affected_stats_combined": affected_stats_combined,
        # Hiển thị số lượng affected_url
        "affected_labels_json": affected_url_stats["affected_labels_json"],
        "affected_total_json": affected_url_stats["affected_total_json"],
        "affected_closed_json": affected_url_stats["affected_closed_json"],

        "exception_stats_combined": exception_stats_combined,

        # Dành cho chart 
        "exception_labels_json": exception_stats["exception_labels_json"],
        "exception_total_json": exception_stats["exception_total_json"],
        "exception_closed_json": exception_stats["exception_closed_json"],
        "pentest_tasks_json": timeline_stats["pentest_tasks_json"],
        "verify_tasks_json": timeline_stats["verify_tasks_json"],
        "user_choices": timeline_stats["user_choices"],

        "top_10_common_vulns": top_vulns["top_common"],
        "top_10_vuln_by_api": top_vulns["top_api"],

        'apis_starts':apis_starts,

    }

    return render(request, "appsec_task/dashboard.html", context)


# export exception approval


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def export_exception_template(request, appsec_task_id):
    task = get_object_or_404(AppSecTask, id=appsec_task_id)
    exceptions = SecurityException.objects.filter(appsec_task=task)

    template_path = os.path.join(settings.BASE_DIR, "media/templates/exception", "template_exception_approval.xlsx")
    wb = load_workbook(template_path)
    ws = wb["Template"]

    # Bổ sung màu chữ theo mức độ
    risk_text_colors = {
        "Critical": "8B0000",  # Đỏ đậm
        "High": "FF0000",      # Đỏ
        "Medium": "FFA500",    # Cam
        "Low": "008000",       # Xanh lá đậm
        "Recommend": "32CD32"  # Xanh lá nhạt
    }

    # ========================
    # ✅ Thay {{domain}}, {{exception_create}} toàn file
    # ========================
    exception1 = SecurityException.objects.filter(appsec_task=task).first()
    base_context = {
        "domain": str(task.environment_prod) if task.environment_prod else "",
        "exception_create":exception1.exception_create.strftime('%d-%b-%y') if exception1.exception_date else "",
    }

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for key, val in base_context.items():
                    if f"{{{{{key}}}}}" in cell.value:
                        cell.value = cell.value.replace(f"{{{{{key}}}}}", val)

    # ========================
    # ✅ Tìm dòng mẫu chứa {{exc.xxx}}
    # ========================
    template_row = None
    for row in ws.iter_rows(min_row=1, max_row=50):
        if any(cell.value and "{{exc." in str(cell.value) for cell in row):
            template_row = row
            break

    if not template_row:
        return HttpResponse("Không tìm thấy dòng mẫu chứa {{exc.xxx}} trong file Excel", status=400)

    start_row = template_row[0].row + 1

    for idx, exception in enumerate(exceptions, start=1):
        context = {
            **base_context,
            "No": idx,
            "exception_create":exception.exception_create,
            "vulnerability": exception.vulnerability or "",
            "overview": exception.overview or "",
            "exploitability": exception.exploitability or "",
            "exploitability_level": exception.exploitability_level or "",
            "impact": exception.impact or "",
            "impact_level": exception.impact_level or "",
            "risk": exception.risk or "",
            "risk_level": exception.risk_level or "",
            "remediation": exception.remediation or "",
            "reason_of_exception": exception.reason_of_exception or "",
            "exception_date": exception.exception_date.strftime('%d-%b-%y') if exception.exception_date else "",
        }

        # Insert dòng mới
        insert_row = start_row + idx - 1
        ws.insert_rows(insert_row)

        # Copy từng cell từ dòng template và thay thế biến
        for col_idx, cell in enumerate(template_row):
            new_cell = ws.cell(row=insert_row, column=col_idx + 1)
            val = cell.value

            if isinstance(val, str) and "{{exc." in val:
                for key, val_replace in context.items():
                    val = val.replace(f"{{{{exc.{key}}}}}", str(val_replace))
                new_cell.value = val
            else:
                new_cell.value = val

        #apply theo màu, nhưng lại màu của cả cell, chứ k tách được
        # for col_idx, cell_template in enumerate(template_row):
        #     new_cell = ws.cell(row=insert_row, column=col_idx + 1)
        #     original_val = cell_template.value
        #     val = original_val

        #     if isinstance(original_val, str) and "{{exc." in original_val:
        #         for key, val_replace in context.items():
        #             val = val.replace(f"{{{{exc.{key}}}}}", str(val_replace))
        #         new_cell.value = val

        #         # ✅ Tô màu chữ và in đậm nếu là risk/impact/exploitability level
        #         font_color = None
        #         font_bold = False

        #         if "{{exc.risk_level}}" in original_val:
        #             level = context["risk_level"]
        #             font_color = risk_text_colors.get(level)
        #             font_bold = True
        #         elif "{{exc.impact_level}}" in original_val:
        #             level = context["impact_level"]
        #             font_color = risk_text_colors.get(level)
        #             font_bold = True
        #         elif "{{exc.exploitability_level}}" in original_val:
        #             level = context["exploitability_level"]
        #             font_color = risk_text_colors.get(level)
        #             font_bold = True

        #         if font_color:
        #             new_cell.font = Font(color=font_color, bold=font_bold)

        #     else:
        #         new_cell.value = val

    # Xoá dòng mẫu
    ws.delete_rows(template_row[0].row)

    # Ghi ra file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Exception_Approval_{task.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def send_outlook_email(subject, message, recipient_list, html_message=None):
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipient_list,
            html_message=html_message,
            fail_silently=False
        )
        return True
    except Exception as e:
        print(f"❌ Gửi mail thất bại: {e}")
        return False


def send_assigned_mail_and_notification(task, username, task_type):
    Notification.objects.create(
        user=username,
        title=f"New {task_type} Task Assigned",
        description=f"You are assigned to {task_type} task: {task.name}",
        url=f"/{task_type}/view/{task.id}",
    )

    html_content = render_to_string("emails/assigned_task.html", {
        "username": username,
        "task_name": task.name,
        "description":f"You are assigned to {task_type} task:",
        "task_url": f"{settings.SERVER_LOCATION}/pentest/view/{task.id}",
    })

    send_outlook_email(
        subject=f"AppSecTool - New {task_type} Task Assigned".upper(),
        message=f"You are assigned to task: {task.name}",
        recipient_list=[f"{username}@fpt.com"],
        html_message=html_content
    )


@login_required
@require_groups(['Pentester', 'Leader', 'Manager'])
def send_reminder_emails(request):
    pentest_tasks = PentestTask.objects.filter(
        Q(status__in=['Not Start', 'In Progress', 'Done']) &
        (Q(start_date__isnull=True) | Q(end_date__isnull=True)),
        is_active=True
    )

    verify_tasks = VerifyTask.objects.filter(
        Q(status__in=['Not Start', 'In Progress', 'Done']) &
        (Q(start_date__isnull=True) | Q(end_date__isnull=True)),
        is_active=True
    )

    # Lấy danh sách user trong hệ thống, convert thành lowercase set
    valid_users = set(user.username.lower() for user in User.objects.all())

    # Key sẽ là username dạng lowercase, Value là dict chứa tasks
    tasks_by_pic = defaultdict(lambda: {'pentest': [], 'verify': []})

    for task in list(pentest_tasks) + list(verify_tasks):
        # Tách PIC_ISM theo dấu `,`, bỏ khoảng trắng và lowercase
        raw_pics = [p.strip().lower() for p in task.PIC_ISM.split(",") if p.strip()]
        for pic in raw_pics:
            if pic in valid_users:
                if isinstance(task, PentestTask):
                    tasks_by_pic[pic]['pentest'].append(task)
                elif isinstance(task, VerifyTask):
                    tasks_by_pic[pic]['verify'].append(task)

    # Truyền tasks_by_pic cho template (pic dạng lowercase)
    return render(request, "appsec_task/tmp_reminder.html", {
        'tasks_by_pic': tasks_by_pic.items(),
        'server_location': settings.SERVER_LOCATION,
    })

